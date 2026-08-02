"""Excel 配置模板：生成空白模板、读取用户填写的内容、转成 YAML。

为什么用 Excel 而不是让用户直接写 YAML：目标用户是医生，日常工作介质就是 Excel。
YAML 的缩进和中文冒号是高频雷区，而这些错误在 Excel 里根本不存在。

分工：**Excel 是给人填的界面，YAML 是给程序读的格式。**
能改 YAML 的高级用户直接改 YAML，两者不冲突。
"""
from __future__ import annotations

import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:                                    # pragma: no cover
    raise SystemExit("缺少依赖 openpyxl，请先运行：pip3 install -r requirements.txt")

from .config import MATCHERS, ConfigError, _VALID_PUB_TYPES

SHEETS = ("设置", "板块", "子板块", "期刊")
GUIDE = "说明"          # 首页说明，不参与解析

_HDR_FILL = PatternFill("solid", fgColor="2F5597")
_NOTE_FILL = PatternFill("solid", fgColor="FFF6DA")
_HDR_FONT = Font(bold=True, size=11, color="FFFFFF", scheme="minor")
_NOTE_FONT = Font(size=9, color="8A6D1B", scheme="minor")

# 下拉选项：与 config.py 的取值一一对应，用户不用记也不会拼错
_DROPDOWNS: dict[str, dict[str, list[str]]] = {
    "设置": {"翻译标题": ["是", "否"],
             # 选项里不能出现逗号：Excel 的下拉列表就是用逗号分隔选项的，
             # 写成 "Q1,Q2" 会被拆成两项。故用 + 连接。
             "允许分区": ["仅Q1", "Q1+Q2", "Q1+Q2+Q3", "不限"]},
    "板块": {"匹配方式": ["关键词", "交叉"],
             "检索范围": ["期刊", "全库"],
             "检索字段": ["标题", "标题+摘要"],
             "质量过滤": ["是", "否"]},
    "期刊": {"收录方式": ["筛选", "全量"]},
}

# 每张表的列定义：(列名, 列宽, 填写说明)
_COLS: dict[str, list[tuple[str, int, str]]] = {
    "设置": [
        ("项目名", 24, "报告标题"),
        ("输出目录", 14, "留空＝output"),
        ("翻译标题", 11, "需配 DEEPL_API_KEY"),
        ("允许分区", 14, "↓这两列决定「质量过滤」按什么筛"),
        ("IF下限", 10, "如 2 表示 IF≥2"),
    ],
    "板块": [
        ("顺序", 7, "去重优先级，小的优先"),
        ("板块名", 18, "如「药物不良反应」"),
        ("匹配方式", 13, "关键词＝只填子板块；交叉＝还要填触发词。详见「说明」页"),
        ("检索范围", 13, "期刊＝只搜期刊表；全库＝搜整个 PubMed"),
        ("检索字段", 13, "留空＝标题"),
        ("质量过滤", 11, "按「设置」页的分区/IF 筛，仅「全库」有效"),
        ("触发词", 38, "只有「交叉」才填；选「关键词」时留空"),
        ("排除词", 26, "标题命中即排除，可留空"),
        ("保留类型", 38, "留空＝不限"),
        ("排除类型", 26, "常填 Editorial,Letter,Comment,Erratum"),
        ("兜底子板块", 14, "配合「全量」收录的刊，可留空"),
    ],
    "子板块": [
        ("所属板块", 18, "须与「板块」页的板块名完全一致"),
        ("顺序", 7, "归类优先级，具体的写前面"),
        ("子板块名", 16, "如「肾脏」"),
        ("关键词", 62, "两种匹配方式都要填。逗号/换行分隔，只写单数原形"),
    ],
    "期刊": [
        ("期刊名", 52, "PubMed 全名，可用 validate 命令核对"),
        ("收录方式", 13, "筛选＝只收命中关键词的（多数用这个）"),
        ("顺序", 7, "报告里的展示顺序，可留空"),
    ],
}

_GUIDE_TEXT = [
    ("h", "怎么填这份表"),
    ("p", "填完保存，然后运行：  python3 cli.py from-excel --excel 本文件.xlsx"),
    ("p", "填错不要紧——转换时会告诉你第几页、第几行、哪一列填错了，以及可以填什么。"),
    ("", ""),
    ("h", "四张表分别管什么"),
    ("p", "设置　　：项目名、是否翻译标题，以及「质量过滤」用的分区与 IF 门槛"),
    ("p", "板块　　：你要分几大类，每类怎么匹配。顺序＝去重优先级"),
    ("p", "子板块　：每个板块下面再分小类，写各自的关键词。顺序＝归类优先级"),
    ("p", "期刊　　：要盯哪些刊。绝大多数填「筛选」即可"),
    ("", ""),
    ("h", "「关键词」和「交叉」的区别——这是最需要想清楚的一项"),
    ("p", "两者的差别只有一句话：**要不要同时满足两类条件**。"),
    ("", ""),
    ("p", "【关键词】只填「子板块」页的关键词就行，板块页的「触发词」留空不填。"),
    ("p", "　　　　　标题命中任意一个关键词即收录。适合「我就关注这几个主题」。"),
    ("p", "　　例：子板块「射血分数保留」关键词填  HFpEF"),
    ("p", "　　　　板块页「触发词」← 留空"),
    ("p", "　　　　→ 收　：Prognosis of HFpEF in elderly patients　（含 HFpEF）"),
    ("p", "　　　　→ 不收：Prognosis of heart failure　　　　　　 （不含任何关键词）"),
    ("", ""),
    ("p", "【交叉】要在**两个地方**都填词，标题必须同时命中两边才收："),
    ("p", "　　　　板块页「触发词」填一类，各子板块「关键词」填另一类。"),
    ("p", "　　　　适合「A 类药 × B 类结局」这种主题，能挡掉大量只沾一边的文献。"),
    ("p", "　　例：板块「药物不良反应」触发词填  SGLT2 inhibitor"),
    ("p", "　　　　子板块「肾脏」关键词填　　　  acute kidney injury"),
    ("p", "　　　　→ 收　：Acute kidney injury associated with SGLT2 inhibitors（两边都有）"),
    ("p", "　　　　→ 不收：SGLT2 inhibitors in heart failure　　　　　　　　　 （缺肾脏词）"),
    ("p", "　　　　→ 不收：Acute kidney injury after cardiac surgery　　　　　 （缺药物词）"),
    ("", ""),
    ("p", "一句话对照："),
    ("p", "　　关键词 → 只填「子板块」的关键词，「触发词」留空"),
    ("p", "　　交叉　 → 「触发词」和「子板块」的关键词都要填，缺一不可"),
    ("", ""),
    ("h", "「质量过滤」怎么设"),
    ("p", "板块页的「质量过滤」只是开关（是/否）；到底按什么筛，在「设置」页定："),
    ("p", "　　允许分区：仅Q1 ／ Q1+Q2 ／ Q1+Q2+Q3 ／ 不限（下拉选）"),
    ("p", "　　IF下限　：填数字，如 2 表示只要 IF≥2 的"),
    ("p", "两个条件同时生效。只对「检索范围＝全库」的板块有意义——"),
    ("p", "走「期刊」表的板块，期刊本来就是你自己挑的，不需要再筛。"),
    ("p", "另需先生成 IF 数据，两条路任选其一："),
    ("p", "　　python3 cli.py import-if　　　　　　　　　　← 从第三方仓库下载，授权边界见 README"),
    ("p", "　　python3 cli.py import-if --excel 你的JCR名单.xlsx　← 用单位订阅的名单，更准"),
    ("p", "没生成也能跑，只是不做这层过滤（宁可多收，也不静默漏掉）。"),
    ("", ""),
    ("h", "关键词怎么写"),
    ("p", "· 只写单数原形：写 neuropathy 就够，程序自动匹配 neuropathies"),
    ("p", "· 按完整单词匹配，不会误伤：optic 不会命中 optical"),
    ("p", "· 词组照常写：acute kidney injury"),
    ("p", "· 多个词用逗号或换行分隔都行，中英文逗号都认"),
    ("", ""),
    ("h", "两个顺序的含义"),
    ("p", "板块顺序　：一篇文章只归**第一个**命中的板块，不会重复出现。最专的写前面。"),
    ("p", "子板块顺序：取**第一个**命中的子板块。更具体的写前面——"),
    ("p", "　　　　　　若「心肌炎」和「心脏」并存，「心肌炎」要排在前面，否则会被「心脏」抢走。"),
    ("p", "　　　　　　习惯上把「其他」放最后。"),
]


# 中文用微软雅黑、英文数字用 Times New Roman。
# xlsx 的单元格字体只能是**一种**字体，中英文分开只能靠**主题字体**实现：
# 主题里 latin 管西文、ea / script="Hans" 管中文，单元格则必须使用主题字体
# （即不写死字体名，只标 scheme="minor"），否则会整体套用那一种字体。
FONT_LATIN = "Times New Roman"
FONT_CJK = "微软雅黑"


def _patch_theme(path: Path) -> None:
    """保存后改写主题字体，并去掉默认字体的写死名字（否则空单元格仍是 Calibri）。"""
    import shutil
    import zipfile

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(
            tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/theme/theme1.xml":
                x = data.decode("utf-8")
                x = re.sub(r'<a:latin typeface="[^"]*"', f'<a:latin typeface="{FONT_LATIN}"', x)
                x = re.sub(r'<a:ea typeface="[^"]*"', f'<a:ea typeface="{FONT_CJK}"', x)
                # 简繁中文脚本回退也一并指到雅黑
                x = re.sub(r'(<a:font script="Han[st]" typeface=")[^"]*"',
                           rf'\1{FONT_CJK}"', x)
                data = x.encode("utf-8")
            elif item.filename == "xl/styles.xml":
                # 默认字体写死了 Calibri，会让用户新填的单元格不走主题
                data = data.decode("utf-8").replace(
                    '<font><name val="Calibri"/>', "<font>").encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


def _write_guide(ws):
    """首页说明。用尽量口语的方式解释两种匹配方式和质量过滤。"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 108
    r = 1
    for kind, text in _GUIDE_TEXT:
        c = ws.cell(r, 1, text)
        if kind == "h":
            c.font = Font(bold=True, size=12, color="2F5597", scheme="minor")
            ws.row_dimensions[r].height = 22
        elif kind == "p":
            c.font = Font(size=10.5, scheme="minor")
        c.alignment = Alignment(vertical="center")
        r += 1
    ws.cell(1, 1).font = Font(bold=True, size=15, color="2F5597", scheme="minor")


def write_template(path: str | Path, rows: int = 200) -> Path:
    """生成空白 Excel 模板：首页说明 + 4 张数据表（带下拉选项与示例行）。"""
    from openpyxl.worksheet.datavalidation import DataValidation

    p = Path(path).expanduser()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_guide(wb.create_sheet(GUIDE))

    thin = Side(style="thin", color="C8D2DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    samples: dict[str, list[list]] = {
        "设置": [["示例文献追踪", "output", "是", "Q1+Q2", 2]],
        "板块": [
            [1, "药物不良反应", "交叉", "全库", "标题", "是",
             "SGLT2 inhibitor, GLP-1 receptor agonist", "",
             "Journal Article, Review, Case Reports",
             "Editorial, Letter, Comment, Erratum", ""],
            [2, "心衰进展", "关键词", "期刊", "标题", "否", "", "study protocol",
             "Journal Article, Review", "Editorial, Letter", ""],
        ],
        "子板块": [
            ["药物不良反应", 1, "肾脏", "acute kidney injury, nephropathy"],
            ["药物不良反应", 2, "代谢", "ketoacidosis, hypoglycemia"],
            ["药物不良反应", 3, "其他", "adverse event, safety"],
            ["心衰进展", 1, "射血分数保留", "HFpEF, preserved ejection fraction"],
            ["心衰进展", 2, "其他", ""],
        ],
        "期刊": [
            ["European heart journal", "全量", 1],
            ["The New England journal of medicine", "筛选", 2],
            ["Nature medicine", "筛选", 3],
        ],
    }

    for name in SHEETS:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        cols = _COLS[name]
        for i, (title, width, note) in enumerate(cols, 1):
            h = ws.cell(1, i, title)
            h.font, h.fill, h.border = _HDR_FONT, _HDR_FILL, border
            h.alignment = Alignment(vertical="center", horizontal="center")
            n = ws.cell(2, i, note)
            n.font, n.fill, n.border = _NOTE_FONT, _NOTE_FILL, border
            n.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 34

        for r, row in enumerate(samples.get(name, []), start=3):
            for cidx, val in enumerate(row, 1):
                c = ws.cell(r, cidx, val)
                c.border = border
                c.font = Font(size=10.5, scheme="minor")
                c.alignment = Alignment(vertical="center", wrap_text=(cols[cidx-1][1] > 30))

        # 下拉选项：省得用户记选项、也不会拼错（「交差/交叉」这类错就没了）
        for col_name, opts in _DROPDOWNS.get(name, {}).items():
            idx = next((i for i, c in enumerate(cols, 1) if c[0] == col_name), None)
            if not idx:
                continue
            dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"',
                                allow_blank=True, showDropDown=False)
            dv.error = "请从下拉列表中选择"
            dv.errorTitle = f"「{col_name}」取值不对"
            ws.add_data_validation(dv)
            letter = get_column_letter(idx)
            dv.add(f"{letter}3:{letter}{rows}")

        ws.freeze_panes = "A3"

    wb.active = 0            # 打开时停在「说明」页
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    _patch_theme(p)
    return p


# ─── 读取 ─────────────────────────────────────────────────────────────────────

_YES = {"是", "y", "yes", "true", "1", "√", "✓", "是的", "要"}
_NO = {"否", "n", "no", "false", "0", "×", "不", "不要"}
_MATCHER_ZH = {"关键词": "simple_keyword", "交叉": "cross_product"}
_SCOPE_ZH = {"期刊": "journals", "全库": "pubmed_all"}
_FIELD_ZH = {"标题": "ti", "标题+摘要": "tiab", "标题和摘要": "tiab"}
_INCLUSION_ZH = {"全量": "full", "筛选": "filtered"}


def _split(val) -> list[str]:
    """把「a, b\nc」这类单元格拆成列表。逗号（中英文）、分号、换行都算分隔符。"""
    if val is None:
        return []
    parts = re.split(r"[,，;；\n]+", str(val))
    return [p.strip() for p in parts if p.strip()]


def _bool(val, default=False, *, where: str = "", field: str = "") -> bool:
    """是/否 解析。**无法识别时报错而不是当成「否」**——静默误解是最糟的失败模式：
    用户填了个表示肯定的词，程序却当否处理，且全程无提示。"""
    if val is None or str(val).strip() == "":
        return default
    v = str(val).strip().lower()
    if v in _YES:
        return True
    if v in _NO:
        return False
    raise ConfigError(
        f"{where or '配置'} 的「{field or '是/否项'}」填了「{val}」，无法识别。\n"
        f"  请填「是」或「否」（下拉里可以直接选）")


def _rows(ws) -> list[dict]:
    """读表：第 1 行列名、第 2 行说明（跳过）、第 3 行起为数据。"""
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    out = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, i).value for i in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        row = {h: v for h, v in zip(headers, vals) if h}
        row["_row"] = r
        out.append(row)
    return out


def _enum(val, mapping: dict, field: str, where: str, default=None):
    s = (str(val).strip() if val is not None else "")
    if not s:
        if default is not None:
            return default
        raise ConfigError(f"{where} 的「{field}」不能为空，可填：{' / '.join(mapping)}")
    if s in mapping:
        return mapping[s]
    if s in mapping.values():          # 也接受直接写英文值
        return s
    raise ConfigError(
        f"{where} 的「{field}」填了「{s}」，无法识别。\n  可填：{' / '.join(mapping)}")


def read_excel(path: str | Path) -> dict:
    """读取填好的 Excel，返回可交给 config 校验的 dict（等价于 YAML 结构）。"""
    p = Path(path).expanduser()
    if not p.exists():
        raise ConfigError(f"Excel 文件不存在：{p}")
    wb = openpyxl.load_workbook(p, data_only=True)

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise ConfigError(
            f"Excel 缺少工作表：{missing}\n"
            f"  需要这 4 张：{'、'.join(SHEETS)}\n"
            f"  建议用 `python3 cli.py template` 重新生成模板后再填。")

    # ── 设置 ──
    st = _rows(wb["设置"])
    s0 = st[0] if st else {}
    out: dict = {
        "project_name": (str(s0.get("项目名")).strip()
                         if s0.get("项目名") else "lit-tracker"),
        "output_dir": (str(s0.get("输出目录")).strip()
                       if s0.get("输出目录") else "output"),
        "translate_titles": _bool(s0.get("翻译标题"), True,
                                  where="「设置」表", field="翻译标题"),
    }
    q = str(s0.get("允许分区") or "").strip()
    if q:
        if q in ("不限", "全部", "不限制"):
            out["allowed_quartiles"] = ["Q1", "Q2", "Q3", "Q4"]
        else:
            # 用正则抓 Q1..Q4，这样「仅Q1」「Q1+Q2」「Q1,Q2」「Q1 Q2」都能认
            qs = re.findall(r"[Qq]\s*([1-4])", q)
            if not qs:
                raise ConfigError(
                    f"「设置」表的「允许分区」填了「{q}」，无法识别。\n"
                    f"  请填 仅Q1 / Q1+Q2 / Q1+Q2+Q3 / 不限（或直接写 Q1、Q1,Q2）")
            out["allowed_quartiles"] = [f"Q{n}" for n in dict.fromkeys(qs)]
    if s0.get("IF下限") not in (None, ""):
        try:
            out["min_impact_factor"] = float(s0["IF下限"])
        except (TypeError, ValueError):
            raise ConfigError(f"「设置」表的「IF下限」应为数字，实际是 {s0['IF下限']!r}")

    # ── 期刊 ──
    full_inc: dict[str, str] = {}
    kw_filt: dict[str, str] = {}
    order_pairs: list[tuple[int, str]] = []
    for row in _rows(wb["期刊"]):
        where = f"「期刊」表第 {row['_row']} 行"
        name = str(row.get("期刊名") or "").strip()
        if not name:
            raise ConfigError(f"{where} 的「期刊名」为空")
        mode = _enum(row.get("收录方式"), _INCLUSION_ZH, "收录方式", where, "filtered")
        disp = str(row.get("显示名") or "").strip() or name   # 无该列时直接用全名
        (full_inc if mode == "full" else kw_filt)[name] = disp
        if row.get("顺序") not in (None, ""):
            try:
                order_pairs.append((int(row["顺序"]), disp))
            except (TypeError, ValueError):
                raise ConfigError(f"{where} 的「顺序」应为整数，实际是 {row['顺序']!r}")
    if full_inc:
        out["full_inclusion_journals"] = full_inc
    if kw_filt:
        out["keyword_filtered_journals"] = kw_filt
    if order_pairs:
        out["journal_order"] = [d for _, d in sorted(order_pairs)]

    # ── 子板块（先读，按所属板块归组）──
    subs_by_sec: dict[str, list[tuple[int, str, list[str]]]] = {}
    for row in _rows(wb["子板块"]):
        where = f"「子板块」表第 {row['_row']} 行"
        sec = str(row.get("所属板块") or "").strip()
        name = str(row.get("子板块名") or "").strip()
        if not sec:
            raise ConfigError(f"{where} 的「所属板块」为空")
        if not name:
            raise ConfigError(f"{where} 的「子板块名」为空")
        try:
            order = int(row["顺序"]) if row.get("顺序") not in (None, "") else 999
        except (TypeError, ValueError):
            raise ConfigError(f"{where} 的「顺序」应为整数，实际是 {row['顺序']!r}")
        subs_by_sec.setdefault(sec, []).append((order, name, _split(row.get("关键词"))))

    # ── 板块 ──
    sec_rows = _rows(wb["板块"])
    if not sec_rows:
        raise ConfigError("「板块」表里没有任何板块（第 3 行起填写；第 2 行是说明，不要删）")

    sections = []
    for row in sec_rows:
        where = f"「板块」表第 {row['_row']} 行"
        name = str(row.get("板块名") or "").strip()
        if not name:
            raise ConfigError(f"{where} 的「板块名」为空")
        try:
            order = int(row["顺序"]) if row.get("顺序") not in (None, "") else 999
        except (TypeError, ValueError):
            raise ConfigError(f"{where} 的「顺序」应为整数，实际是 {row['顺序']!r}")

        matcher = _enum(row.get("匹配方式"), _MATCHER_ZH, "匹配方式", where, "simple_keyword")
        scope = _enum(row.get("检索范围"), _SCOPE_ZH, "检索范围", where, "journals")
        field = _enum(row.get("检索字段"), _FIELD_ZH, "检索字段", where, "ti")

        if name not in subs_by_sec:
            raise ConfigError(
                f"板块「{name}」在「子板块」表里没有任何子板块。\n"
                f"  请在「子板块」表新增行，「所属板块」填「{name}」。")

        subs = []
        for _, sname, kws in sorted(subs_by_sec[name]):
            sub: dict = {"name": sname}
            if matcher == "cross_product":
                sub["cross_keywords"] = kws
            else:
                sub["keywords"] = kws
            subs.append(sub)

        sec: dict = {"name": name, "matcher": matcher, "scope": scope,
                     "search_field": field, "subsections": subs}
        if _bool(row.get("质量过滤"), where=where, field="质量过滤"):
            sec["quality_filter"] = True
        for key, col in (("trigger_keywords", "触发词"), ("exclude_keywords", "排除词"),
                         ("include_types", "保留类型"), ("exclude_types", "排除类型")):
            vals = _split(row.get(col))
            if vals:
                sec[key] = vals
        fb = str(row.get("兜底子板块") or "").strip()
        if fb:
            sec["fallback_subsection"] = fb
        sections.append((order, sec))

    # 「子板块」表里写了、但「板块」表里没有的板块名 —— 最常见的填写错误
    known = {str(r.get("板块名") or "").strip() for r in sec_rows}
    orphan = sorted(set(subs_by_sec) - known)
    if orphan:
        raise ConfigError(
            f"「子板块」表里这些「所属板块」在「板块」表里找不到：{orphan}\n"
            f"  「板块」表现有：{sorted(known)}\n"
            f"  多半是名字写得不完全一致（多空格、错别字）。")

    out["sections"] = [s for _, s in sorted(sections, key=lambda x: x[0])]
    return out


def to_yaml(data: dict, path: str | Path) -> Path:
    """把读出的结构写成 YAML（供程序使用，也便于高级用户继续手改）。"""
    import yaml
    p = Path(path).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)
    header = ("# 本文件由 Excel 模板自动生成（cli.py from-excel）。\n"
              "# 可直接手改；若之后重新从 Excel 生成，手改内容会被覆盖。\n\n")
    body = yaml.safe_dump(data, allow_unicode=True, sort_keys=False, width=100)
    p.write_text(header + body, encoding="utf-8")
    return p
