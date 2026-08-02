"""生成 if_data.json（期刊 IF 与 JCR 分区）。两条路，任选其一：

  ① download —— 从第三方公开仓库（EasyPubMed）下载其整理好的数据包。
     省事，但可能略滞后于官方 JCR。

  ② from-excel —— 用自己单位订阅的 JCR 名单 Excel 导入，数据最新最准。

**版权边界（重要）**：本项目不分发 if_data.json，也不对其中数据主张任何权利。
IF/分区的原始指标出自 Clarivate 的 Journal Citation Reports，属商业订阅产品，
其使用与再分发受 Clarivate 条款约束（https://clarivate.com/legal/terms-of-use/）。
①中 EasyPubMed 仓库自身采用 MIT，但**仓库的许可证不会让其中收录的第三方数据
自动变成 MIT 数据**；本项目未核实该数据包的独立授权，走①即由用户自行判断
其在自己场景下的可用性。对授权有顾虑的，请用②导入本单位订阅的名单。
"""
from __future__ import annotations

import datetime
import io
import json
import re
import zipfile
from pathlib import Path

# JCR 标准导出的列位置（0 基）
C_JOURNAL, C_ABBR, C_ISSN, C_EISSN, C_JIF, C_QUARTILE = 0, 1, 3, 4, 8, 9


def _norm_issn(s) -> str:
    return re.sub(r"[^0-9Xx]", "", str(s or "")).upper()[:8]


def _norm_name(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def build(excel_path: str | Path, out_path: str | Path,
          *, merge: bool = True) -> dict:
    """读 JCR Excel，生成/合并 if_data.json。返回统计信息。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("缺少 openpyxl，请先运行：pip3 install -r requirements.txt")

    src = Path(excel_path).expanduser()
    if not src.exists():
        raise SystemExit(f"找不到文件：{src}")
    out = Path(out_path).expanduser()

    data: dict = {"by_issn": {}, "by_name": {}}
    if merge and out.exists():
        try:
            old = json.loads(out.read_text(encoding="utf-8"))
            data["by_issn"] = old.get("by_issn", {})
            data["by_name"] = old.get("by_name", {})
        except Exception:
            pass                                  # 旧文件坏了就重建，不阻塞

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    n_read = n_ok = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) <= C_QUARTILE:
            continue
        journal, jif_raw = row[C_JOURNAL], row[C_JIF]
        if not journal or jif_raw in (None, ""):
            continue
        n_read += 1
        try:
            jif = str(round(float(jif_raw), 1))
        except (TypeError, ValueError):
            continue                              # 「N/A」等非数值，跳过
        q = str(row[C_QUARTILE]).strip() if row[C_QUARTILE] else ""
        rec = {"if": jif, "quartile": q}
        for name in (journal, row[C_ABBR]):
            for k in {_norm_name(name), _norm_loose(name)}:
                if k:
                    data["by_name"][k] = rec
        for issn in (row[C_ISSN], row[C_EISSN]):
            k = _norm_issn(issn)
            if len(k) == 8:
                data["by_issn"][k] = rec
        n_ok += 1
    wb.close()

    data["generated"] = datetime.date.today().isoformat()
    data["source"] = src.name
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    return {"read": n_read, "ok": n_ok,
            "issn": len(data["by_issn"]), "name": len(data["by_name"]), "path": out}


# ─── 方式二：从第三方公开仓库下载（授权边界见模块开头）──────────────────────
_GITHUB_API = "https://api.github.com/repos/naivenaive/EasyPubMed/contents/"
_DATA_FILE = "jcr_cas_ifqb.json"


def _norm_loose(name: str) -> str:
    """宽松键：去掉独立的 and / the，兜住「A and B」vs「A & B」、
    「The Journal of X」vs「Journal of X」这类写法差异。
    必须与 journals.py 的 _norm_loose 保持一致，否则存进去也查不到。"""
    s = re.sub(r"[^\w\s]", " ", (name or "").lower())
    s = re.sub(r"\b(and|the)\b", " ", s)
    return re.sub(r"[^a-z0-9]", "", s)


def download(out_path: str | Path, *, merge: bool = False) -> dict:
    """从 EasyPubMed 仓库下载其整理的 JCR 数据，生成 if_data.json。

    注意授权边界（模块开头有详述）：数据源的许可证状态未经核实，
    原始指标版权归 Clarivate。
    """
    import requests

    out = Path(out_path).expanduser()
    r = requests.get(_GITHUB_API, timeout=30,
                     headers={"Accept": "application/vnd.github.v3+json"})
    r.raise_for_status()
    zip_url = zip_name = None
    for item in r.json():
        if item["name"].lower().endswith(".zip") and "easypubmed" in item["name"].lower():
            zip_url, zip_name = item["download_url"], item["name"]
            break
    if not zip_url:
        raise RuntimeError(
            "未在该仓库根目录找到数据包，仓库结构可能已变。\n"
            "  可改用：python3 cli.py import-if --excel <你的JCR名单.xlsx>")

    rz = requests.get(zip_url, timeout=180)
    rz.raise_for_status()
    z = zipfile.ZipFile(io.BytesIO(rz.content))
    hits = [n for n in z.namelist() if n.endswith(_DATA_FILE)]
    if not hits:
        raise RuntimeError(f"数据包里没有 {_DATA_FILE}，仓库结构可能已变。")
    entries = json.loads(z.read(hits[0]))

    data: dict = {"by_issn": {}, "by_name": {}}
    if merge and out.exists():
        try:
            old = json.loads(out.read_text(encoding="utf-8"))
            data["by_issn"] = old.get("by_issn", {})
            data["by_name"] = old.get("by_name", {})
        except Exception:
            pass

    n = 0
    for e in entries:
        jif = (e.get("IF") or "").strip()
        q = (e.get("Q") or "").strip()
        if not jif or jif == "N/A":
            continue
        rec = {"if": jif, "quartile": q if q and q != "N/A" else ""}
        for key in ("issn", "eissn"):
            raw = (e.get(key) or "").strip()
            if raw and raw != "N/A":
                data["by_issn"].setdefault(raw.replace("-", ""), rec)
        for key in ("journal", "jabb"):
            raw = e.get(key) or ""
            # 严格键供精确匹配，宽松键兜住 and/& 的写法差异，两者都存
            for k in {_norm_name(raw), _norm_loose(raw)}:
                if k:
                    data["by_name"].setdefault(k, rec)
        n += 1

    data["generated"] = datetime.date.today().isoformat()
    data["source"] = f"EasyPubMed 仓库整理的 JCR 数据（{zip_name}）"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    return {"read": len(entries), "ok": n,
            "issn": len(data["by_issn"]), "name": len(data["by_name"]), "path": out}
