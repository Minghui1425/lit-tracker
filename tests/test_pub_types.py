"""文章类型：清单完整性、拼写纠正、带逗号类型的切分。

类型名写错**不会报错，只会静默筛不到东西**——最糟的失败模式。
所以这里的重点是「写错时能不能被拦下并给出正确写法」。
"""
import pytest

from littrack import config as cfgmod
from littrack import excel


def test_table_is_consistent():
    names = [t for t, _, _ in cfgmod.PUB_TYPES]
    assert len(names) == len(set(names)), "类型表里有重复"
    assert set(names) == cfgmod._VALID_PUB_TYPES
    assert {a for _, _, a in cfgmod.PUB_TYPES} == {"保留", "排除", "视情况"}
    assert all(zh.strip() for _, zh, _ in cfgmod.PUB_TYPES), "每条都要有中文说明"


def test_table_covers_the_common_ones():
    names = set(t for t, _, _ in cfgmod.PUB_TYPES)
    assert {"Journal Article", "Review", "Systematic Review", "Meta-Analysis",
            "Randomized Controlled Trial", "Case Reports", "Observational Study",
            "Editorial", "Letter", "Comment", "Practice Guideline"} <= names
    assert len(names) >= 50, "总表太短就失去了「不用自己想」的意义"


def test_the_erratum_trap_is_fixed():
    """`Erratum` 不是 PubMed 的类型名（实测 Erratum[pt] 零命中），
    真实写法是 `Published Erratum`。旧版示例配置写错过，这条筛选从未生效。"""
    assert "Erratum" not in cfgmod._VALID_PUB_TYPES
    assert "Published Erratum" in cfgmod._VALID_PUB_TYPES
    assert "Retraction of Publication" not in cfgmod._VALID_PUB_TYPES
    assert {"Retraction Notice", "Retracted Publication"} <= cfgmod._VALID_PUB_TYPES


def test_default_types_split_the_table():
    keep, drop = cfgmod.default_types("保留"), cfgmod.default_types("排除")
    assert "Journal Article" in keep and "Editorial" in drop
    assert not set(keep) & set(drop)


def _cfg(inc=None, exc=None):
    return {"project_name": "t", "keyword_filtered_journals": {"JAMA": "JAMA"},
            "sections": [{"name": "板块甲", "matcher": "simple_keyword", "scope": "journals",
                          "include_types": inc or [], "exclude_types": exc or [],
                          "subsections": [{"name": "其他", "keywords": ["sepsis"]}]}]}


def test_case_is_normalized():
    c = cfgmod.load_dict(_cfg(inc=["journal article", "REVIEW", " Meta-Analysis "]))
    assert c.sections[0].include_types == ["Journal Article", "Review", "Meta-Analysis"]


def test_typo_gets_a_suggestion():
    with pytest.raises(cfgmod.ConfigError) as e:
        cfgmod.load_dict(_cfg(exc=["Erratum"]))
    assert "Published Erratum" in str(e.value)          # 给出正确写法，而不只是报错

    with pytest.raises(cfgmod.ConfigError) as e:
        cfgmod.load_dict(_cfg(inc=["Case Report"]))     # 少了复数 s
    assert "Case Reports" in str(e.value)


def test_phase_trials_survive_yaml_and_excel(tmp_path):
    c = cfgmod.load_dict(_cfg(inc=["Clinical Trial, Phase III"]))
    assert c.sections[0].include_types == ["Clinical Trial, Phase III"]


def test_excel_split_does_not_cut_types_at_their_commas():
    """按逗号硬切会把「Clinical Trial, Phase III」腰斩成「Clinical Trial」——
    后者是个合法但宽得多的类型，筛选范围会被悄悄放大。"""
    got = excel._split_types(
        "Journal Article, Clinical Trial, Phase III, Review")
    assert got == ["Journal Article", "Clinical Trial, Phase III", "Review"]
    assert "Phase III" not in got

    # 换行分隔、大小写不一致，同样要还原成规范写法
    assert excel._split_types("clinical trial, phase i\nReview") == \
           ["Clinical Trial, Phase I", "Review"]


def test_plain_split_still_used_for_keywords():
    assert excel._split("a, b\nc；d") == ["a", "b", "c", "d"]


def test_template_has_a_types_sheet_with_the_whole_table(tmp_path):
    import openpyxl
    p = excel.write_template(tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(p)
    assert excel.TYPES_SHEET in wb.sheetnames
    ws = wb[excel.TYPES_SHEET]
    listed = {r[0] for r in ws.iter_rows(min_row=7, values_only=True) if r[0]}
    assert listed == cfgmod._VALID_PUB_TYPES          # 模板与代码不能漂移


def test_template_prefills_both_type_columns(tmp_path):
    import openpyxl
    p = excel.write_template(tmp_path / "t.xlsx")
    ws = openpyxl.load_workbook(p)["板块"]
    header = [c.value for c in ws[1]]
    inc = ws.cell(3, header.index("保留类型") + 1).value
    exc = ws.cell(3, header.index("排除类型") + 1).value
    assert excel._split_types(inc) == cfgmod.default_types("保留")
    assert excel._split_types(exc) == cfgmod.default_types("排除")


def test_template_round_trip_keeps_every_type(tmp_path):
    p = excel.write_template(tmp_path / "t.xlsx")
    data = excel.read_excel(p)
    cfg = cfgmod.load_dict(data)
    assert cfg.sections[0].include_types == cfgmod.default_types("保留")
    assert cfg.sections[0].exclude_types == cfgmod.default_types("排除")


def test_example_config_uses_valid_types(cfg):
    for s in cfg.sections:
        assert set(s.include_types) <= cfgmod._VALID_PUB_TYPES
        assert set(s.exclude_types) <= cfgmod._VALID_PUB_TYPES
